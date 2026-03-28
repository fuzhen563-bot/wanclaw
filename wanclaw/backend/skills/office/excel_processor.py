"""
Excel处理技能 V2.0
基于 openpyxl 实现真实文件操作
"""

import os
import logging
from typing import Dict, List, Optional, Any
from datetime import datetime

from wanclaw.backend.skills import BaseSkill, SkillResult, SkillCategory, SkillLevel

logger = logging.getLogger(__name__)


class ExcelProcessorSkill(BaseSkill):
    
    def __init__(self):
        super().__init__()
        self.name = "ExcelProcessor"
        self.description = "Excel处理：多表合并、拆分、去重、筛选、排序、汇总，生成日报周报月报"
        self.category = SkillCategory.OFFICE
        self.level = SkillLevel.INTERMEDIATE
        self.required_params = ["action"]
        self.optional_params = {
            "file_path": str, "file_paths": list, "output_path": str,
            "sheet_name": str, "merge_type": str, "key_column": (int, str),
            "filter_column": (int, str), "filter_value": str,
            "sort_column": (int, str), "sort_ascending": bool,
            "group_by": (int, str), "aggregate_column": (int, str),
            "aggregate_func": str, "report_type": str,
            "find_text": str, "replace_text": str, "date_format": str
        }

    async def execute(self, params: Dict[str, Any]) -> SkillResult:
        action = params.get("action", "").lower()
        try:
            method = getattr(self, f"_action_{action}", None)
            if not method:
                return SkillResult(success=False, message=f"不支持的操作: {action}")
            return await method(params)
        except Exception as e:
            logger.error(f"Excel处理失败: {action} - {e}")
            return SkillResult(success=False, message=f"Excel处理失败: {str(e)}", error=str(e))

    async def _load_openpyxl(self):
        try:
            import openpyxl
            return openpyxl
        except ImportError:
            raise ImportError("openpyxl 未安装，请运行: pip install openpyxl")

    async def _read_workbook(self, file_path: str):
        openpyxl = await self._load_openpyxl()
        wb = openpyxl.load_workbook(file_path)
        return wb

    async def _action_merge(self, params: Dict) -> SkillResult:
        file_paths = params.get("file_paths", [])
        merge_type = params.get("merge_type", "vertical")
        output_path = params.get("output_path", "merged_output.xlsx")

        if not file_paths:
            return SkillResult(success=False, message="需要提供文件列表", error="file_paths required")

        openpyxl = await self._load_openpyxl()
        output_wb = openpyxl.Workbook()
        output_ws = output_wb.active
        output_ws.title = "合并数据"
        total_rows = 0

        for fp in file_paths:
            if not os.path.exists(fp):
                continue
            wb = openpyxl.load_workbook(fp)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for row in ws.iter_rows(values_only=True):
                    output_ws.append(row)
                    total_rows += 1
            wb.close()

        output_wb.save(output_path)
        output_wb.close()

        return SkillResult(
            success=True,
            message=f"合并完成，共 {total_rows} 行数据",
            data={"merged_rows": total_rows, "source_files": file_paths,
                  "merge_type": merge_type, "output_file": output_path}
        )

    async def _action_split(self, params: Dict) -> SkillResult:
        file_path = params.get("file_path", "")
        key_column = params.get("key_column", 0)
        output_dir = params.get("output_path", os.path.dirname(file_path) or ".")

        if not file_path or not os.path.exists(file_path):
            return SkillResult(success=False, message="文件不存在", error="file not found")

        openpyxl = await self._load_openpyxl()
        wb = await self._read_workbook(file_path)
        ws = wb.active

        groups: Dict[str, List] = {}
        for row in ws.iter_rows(values_only=True):
            key = row[key_column] if key_column < len(row) else None
            if key is not None:
                key_str = str(key)
                if key_str not in groups:
                    groups[key_str] = []
                groups[key_str].append(row)

        os.makedirs(output_dir, exist_ok=True)
        splits = {}
        for key, rows in groups.items():
            out_wb = openpyxl.Workbook()
            out_ws = out_wb.active
            out_ws.title = key
            for row in rows:
                out_ws.append(row)
            out_file = os.path.join(output_dir, f"split_{key}.xlsx")
            out_wb.save(out_file)
            out_wb.close()
            splits[key] = {"rows": len(rows), "output_file": out_file}

        wb.close()
        return SkillResult(
            success=True, message=f"拆分完成，按第{key_column}列拆分为{len(splits)}个文件",
            data={"splits": splits, "split_column": key_column}
        )

    async def _action_deduplicate(self, params: Dict) -> SkillResult:
        file_path = params.get("file_path", "")
        key_column = params.get("key_column", 0)
        output_path = params.get("output_path", file_path)

        if not file_path or not os.path.exists(file_path):
            return SkillResult(success=False, message="文件不存在", error="file not found")

        openpyxl = await self._load_openpyxl()
        wb = await self._read_workbook(file_path)
        ws = wb.active

        seen = set()
        kept = []
        removed = 0
        for row in ws.iter_rows(values_only=True):
            key = row[key_column] if key_column < len(row) else None
            if key not in seen:
                seen.add(key)
                kept.append(row)
            else:
                removed += 1

        out_wb = openpyxl.Workbook()
        out_ws = out_wb.active
        for row in kept:
            out_ws.append(row)
        out_wb.save(output_path)
        out_wb.close()
        wb.close()

        return SkillResult(
            success=True, message=f"去重完成，移除{removed}条重复记录",
            data={"original_rows": removed + len(kept), "duplicate_removed": removed,
                  "remaining_rows": len(kept), "key_column": key_column,
                  "output_file": output_path}
        )

    async def _action_filter(self, params: Dict) -> SkillResult:
        file_path = params.get("file_path", "")
        filter_column = params.get("filter_column", 0)
        filter_value = params.get("filter_value", "")
        output_path = params.get("output_path", file_path.replace(".xlsx", "_filtered.xlsx"))

        if not file_path or not os.path.exists(file_path):
            return SkillResult(success=False, message="文件不存在", error="file not found")

        openpyxl = await self._load_openpyxl()
        wb = await self._read_workbook(file_path)
        ws = wb.active

        matched = []
        for row in ws.iter_rows(values_only=True):
            if filter_column < len(row) and filter_value.lower() in str(row[filter_column]).lower():
                matched.append(row)

        out_wb = openpyxl.Workbook()
        out_ws = out_wb.active
        for row in matched:
            out_ws.append(row)
        out_wb.save(output_path)
        out_wb.close()
        wb.close()

        return SkillResult(
            success=True, message=f"筛选完成，符合条件的有{len(matched)}行",
            data={"filter_column": filter_column, "filter_value": filter_value,
                  "matched_rows": len(matched), "output_file": output_path}
        )

    async def _action_sort(self, params: Dict) -> SkillResult:
        file_path = params.get("file_path", "")
        sort_column = params.get("sort_column", 0)
        sort_ascending = params.get("sort_ascending", True)
        output_path = params.get("output_path", file_path.replace(".xlsx", "_sorted.xlsx"))

        if not file_path or not os.path.exists(file_path):
            return SkillResult(success=False, message="文件不存在", error="file not found")

        openpyxl = await self._load_openpyxl()
        wb = await self._read_workbook(file_path)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        if rows:
            header = rows[0]
            data = rows[1:]
            data.sort(key=lambda r: r[sort_column] if sort_column < len(r) else "", reverse=not sort_ascending)
            rows = [header] + data

        out_wb = openpyxl.Workbook()
        out_ws = out_wb.active
        for row in rows:
            out_ws.append(row)
        out_wb.save(output_path)
        out_wb.close()
        wb.close()

        return SkillResult(
            success=True,
            message=f"排序完成，按第{sort_column}列{'升序' if sort_ascending else '降序'}排列",
            data={"sorted_rows": len(rows) - 1, "sort_column": sort_column,
                  "sort_ascending": sort_ascending, "output_file": output_path}
        )

    async def _action_aggregate(self, params: Dict) -> SkillResult:
        file_path = params.get("file_path", "")
        group_by = params.get("group_by", 1)
        aggregate_func = params.get("aggregate_func", "sum")
        aggregate_column = params.get("aggregate_column", 2)

        if not file_path or not os.path.exists(file_path):
            return SkillResult(success=False, message="文件不存在", error="file not found")

        openpyxl = await self._load_openpyxl()
        wb = await self._read_workbook(file_path)
        ws = wb.active

        groups: Dict[str, List] = {}
        for row in ws.iter_rows(values_only=True):
            if len(row) <= max(group_by, aggregate_column):
                continue
            key = str(row[group_by])
            val = row[aggregate_column]
            try:
                val = float(val)
            except (TypeError, ValueError):
                val = 0
            if key not in groups:
                groups[key] = []
            groups[key].append(val)

        summary = [["分组", "汇总值", "记录数"]]
        for key, vals in groups.items():
            if aggregate_func == "sum":
                agg_val = sum(vals)
            elif aggregate_func == "avg":
                agg_val = sum(vals) / len(vals) if vals else 0
            elif aggregate_func == "max":
                agg_val = max(vals)
            elif aggregate_func == "min":
                agg_val = min(vals)
            else:
                agg_val = len(vals)
            summary.append([key, round(agg_val, 2), len(vals)])

        wb.close()

        out_wb = openpyxl.Workbook()
        out_ws = out_wb.active
        for row in summary:
            out_ws.append(row)
        output_path = params.get("output_path", file_path.replace(".xlsx", "_summary.xlsx"))
        out_wb.save(output_path)
        out_wb.close()

        return SkillResult(
            success=True, message="汇总完成",
            data={"summary": summary, "group_by_column": group_by,
                  "aggregate_function": aggregate_func, "output_file": output_path}
        )

    async def _action_report(self, params: Dict) -> SkillResult:
        file_path = params.get("file_path", "")
        report_type = params.get("report_type", "daily")
        output_path = params.get("output_path", file_path.replace(".xlsx", f"_{report_type}_report.xlsx"))

        if not file_path or not os.path.exists(file_path):
            return SkillResult(success=False, message="文件不存在", error="file not found")

        openpyxl = await self._load_openpyxl()
        wb = await self._read_workbook(file_path)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        out_wb = openpyxl.Workbook()
        out_ws = out_wb.active
        out_ws.title = f"{report_type}报告"

        title = f"{report_type.title()} 销售报告"
        out_ws.append([title])
        out_ws.append(["生成时间:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        out_ws.append([])
        out_ws.append(["=" * 60])
        out_ws.append([])

        if rows:
            out_ws.append(rows[0])
            for row in rows[1:]:
                out_ws.append(row)

            numeric_cols = []
            for i, val in enumerate(rows[0]):
                try:
                    sum_val = sum(float(r[i]) for r in rows[1:] if i < len(r) and r[i] is not None)
                    numeric_cols.append((i, val, sum_val))
                except (TypeError, ValueError, IndexError):
                    pass

            out_ws.append([])
            out_ws.append(["=" * 60])
            out_ws.append(["汇总:"])
            for col_idx, col_name, col_sum in numeric_cols:
                out_ws.append([f"{col_name} 合计", round(col_sum, 2)])

        out_wb.save(output_path)
        out_wb.close()

        return SkillResult(
            success=True, message=f"{report_type}报告生成完成",
            data={"report_type": report_type, "output_file": output_path,
                  "generated_at": datetime.now().isoformat(), "total_rows": len(rows)}
        )

    async def _action_replace(self, params: Dict) -> SkillResult:
        file_path = params.get("file_path", "")
        find_text = params.get("find_text", "")
        replace_text = params.get("replace_text", "")
        output_path = params.get("output_path", file_path.replace(".xlsx", "_replaced.xlsx"))

        if not file_path or not os.path.exists(file_path):
            return SkillResult(success=False, message="文件不存在", error="file not found")

        openpyxl = await self._load_openpyxl()
        wb = await self._read_workbook(file_path)
        replacements = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and find_text in cell.value:
                        cell.value = cell.value.replace(find_text, replace_text)
                        replacements += 1

        wb.save(output_path)
        wb.close()

        return SkillResult(
            success=True, message=f"批量替换完成，替换了{replacements}处",
            data={"find_text": find_text, "replace_text": replace_text,
                  "replacements_made": replacements, "output_file": output_path}
        )
