"""
表格处理技能
提供Excel/CSV文件处理功能
"""

import os
import csv
import json
import logging
import tempfile
from typing import Dict, List, Optional, Any, Tuple
from datetime import datetime

from wanclaw.backend.skills import BaseSkill, SkillResult, SkillCategory, SkillLevel
from wanclaw.backend.im_adapter.security import get_security, OperationType


logger = logging.getLogger(__name__)


class SpreadsheetHandlerSkill(BaseSkill):
    """表格处理技能"""
    
    def __init__(self):
        super().__init__()
        self.name = "SpreadsheetHandler"
        self.description = "表格处理：读取、写入、分析Excel/CSV文件"
        self.category = SkillCategory.OFFICE
        self.level = SkillLevel.INTERMEDIATE
        
        # 必需参数
        self.required_params = ["action", "file_path"]
        
        # 可选参数及其类型
        self.optional_params = {
            "sheet_name": str,
            "data": list,
            "headers": list,
            "start_row": int,
            "end_row": int,
            "start_col": int,
            "end_col": int,
            "search_value": str,
            "filter_column": int,
            "filter_value": str,
            "sort_column": int,
            "sort_ascending": bool,
            "output_path": str,
            "delimiter": str,
            "quotechar": str,
            "format": str
        }
    
    async def execute(self, params: Dict[str, Any]) -> SkillResult:
        """
        执行表格处理操作
        
        Args:
            params: {
                "action": "read|write|info|search|filter|sort|convert|stats",
                "file_path": "文件路径",
                "sheet_name": "工作表名称",
                "data": [["列1", "列2"], ["值1", "值2"]],
                "headers": ["列1", "列2"],
                "start_row": 0,
                "end_row": 100,
                "start_col": 0,
                "end_col": 10,
                "search_value": "搜索值",
                "filter_column": 0,
                "filter_value": "过滤值",
                "sort_column": 0,
                "sort_ascending": true,
                "output_path": "输出路径",
                "delimiter": ",",
                "quotechar": "\"",
                "format": "excel|csv|json"
            }
            
        Returns:
            执行结果
        """
        action = params.get("action", "").lower()
        file_path = params.get("file_path", "")
        user_id = params.get("user_id", "unknown")
        username = params.get("username", "unknown")
        
        # 安全检查
        security = get_security()
        
        if action in ["read", "info", "search", "filter", "sort", "stats", "convert"]:
            # 读取操作
            allowed, reason = security.check_file_access(
                file_path, OperationType.FILE_READ, user_id, username
            )
            if not allowed:
                return SkillResult(
                    success=False,
                    message=f"文件访问被拒绝: {reason}",
                    error="Security check failed"
                )
        elif action == "write":
            # 写入操作
            allowed, reason = security.check_file_access(
                file_path, OperationType.FILE_WRITE, user_id, username
            )
            if not allowed:
                return SkillResult(
                    success=False,
                    message=f"文件写入被拒绝: {reason}",
                    error="Security check failed"
                )
        else:
            return SkillResult(
                success=False,
                message=f"不支持的操作: {action}",
                error=f"Unsupported action: {action}"
            )
        
        # 检查文件扩展名
        file_ext = os.path.splitext(file_path)[1].lower()
        is_excel = file_ext in ['.xlsx', '.xls', '.xlsm']
        is_csv = file_ext == '.csv'
        
        if not is_excel and not is_csv:
            return SkillResult(
                success=False,
                message=f"不支持的文件格式: {file_ext}",
                error=f"Unsupported file format: {file_ext}"
            )
        
        # 执行具体操作
        try:
            if action == "read":
                return await self._read_spreadsheet(params, is_excel)
            elif action == "write":
                return await self._write_spreadsheet(params, is_excel)
            elif action == "info":
                return await self._spreadsheet_info(params, is_excel)
            elif action == "search":
                return await self._search_spreadsheet(params, is_excel)
            elif action == "filter":
                return await self._filter_spreadsheet(params, is_excel)
            elif action == "sort":
                return await self._sort_spreadsheet(params, is_excel)
            elif action == "convert":
                return await self._convert_spreadsheet(params, is_excel)
            elif action == "stats":
                return await self._spreadsheet_stats(params, is_excel)
            else:
                return SkillResult(
                    success=False,
                    message=f"未知操作: {action}",
                    error="Unknown action"
                )
        except Exception as e:
            logger.error(f"表格操作失败: {action} - {e}")
            return SkillResult(
                success=False,
                message=f"表格操作失败: {str(e)}",
                error=str(e)
            )
    
    async def _read_spreadsheet(self, params: Dict[str, Any], is_excel: bool) -> SkillResult:
        """读取表格数据"""
        file_path = params.get("file_path", "")
        sheet_name = params.get("sheet_name", "")
        start_row = params.get("start_row", 0)
        end_row = params.get("end_row", -1)
        start_col = params.get("start_col", 0)
        end_col = params.get("end_col", -1)
        
        if not os.path.exists(file_path):
            return SkillResult(
                success=False,
                message=f"文件不存在: {file_path}",
                error="File not found"
            )
        
        try:
            if is_excel:
                data, sheet_names = self._read_excel(
                    file_path, sheet_name, start_row, end_row, start_col, end_col
                )
            else:
                data = self._read_csv(file_path, start_row, end_row, start_col, end_col)
                sheet_names = ["CSV Data"]
            
            # 限制返回数据量（防止过大响应）
            max_rows = 1000
            if len(data) > max_rows:
                data = data[:max_rows]
                truncated = True
            else:
                truncated = False
            
            return SkillResult(
                success=True,
                message=f"读取 {len(data)} 行数据" + (" (已截断)" if truncated else ""),
                data={
                    "file_path": file_path,
                    "sheet_names": sheet_names,
                    "data": data,
                    "total_rows": len(data),
                    "truncated": truncated,
                    "start_row": start_row,
                    "end_row": end_row if end_row != -1 else "全部",
                    "start_col": start_col,
                    "end_col": end_col if end_col != -1 else "全部"
                }
            )
        except Exception as e:
            return SkillResult(
                success=False,
                message=f"读取表格失败: {str(e)}",
                error=str(e)
            )
    
    async def _write_spreadsheet(self, params: Dict[str, Any], is_excel: bool) -> SkillResult:
        """写入表格数据"""
        file_path = params.get("file_path", "")
        data = params.get("data", [])
        headers = params.get("headers", [])
        sheet_name = params.get("sheet_name", "Sheet1")
        
        if not data:
            return SkillResult(
                success=False,
                message="需要数据",
                error="Data required"
            )
        
        try:
            if is_excel:
                self._write_excel(file_path, data, headers, sheet_name)
            else:
                self._write_csv(file_path, data, headers)
            
            file_size = os.path.getsize(file_path)
            
            return SkillResult(
                success=True,
                message=f"写入 {len(data)} 行数据到 {file_path}",
                data={
                    "file_path": file_path,
                    "rows_written": len(data),
                    "columns_written": len(data[0]) if data else 0,
                    "file_size": file_size,
                    "format": "excel" if is_excel else "csv"
                }
            )
        except Exception as e:
            return SkillResult(
                success=False,
                message=f"写入表格失败: {str(e)}",
                error=str(e)
            )
    
    async def _spreadsheet_info(self, params: Dict[str, Any], is_excel: bool) -> SkillResult:
        """获取表格信息"""
        file_path = params.get("file_path", "")
        
        if not os.path.exists(file_path):
            return SkillResult(
                success=False,
                message=f"文件不存在: {file_path}",
                error="File not found"
            )
        
        try:
            if is_excel:
                sheet_names, row_counts, column_counts = self._get_excel_info(file_path)
                file_type = "Excel"
            else:
                sheet_names, row_counts, column_counts = self._get_csv_info(file_path)
                file_type = "CSV"
            
            file_size = os.path.getsize(file_path)
            stat = os.stat(file_path)
            
            return SkillResult(
                success=True,
                message=f"{file_type} 文件信息",
                data={
                    "file_path": file_path,
                    "file_type": file_type,
                    "file_size": file_size,
                    "size_formatted": self._format_size(file_size),
                    "modified": datetime.fromtimestamp(stat.st_mtime).isoformat(),
                    "created": datetime.fromtimestamp(stat.st_ctime).isoformat(),
                    "sheets": [
                        {
                            "name": sheet_names[i],
                            "rows": row_counts[i],
                            "columns": column_counts[i]
                        }
                        for i in range(len(sheet_names))
                    ]
                }
            )
        except Exception as e:
            return SkillResult(
                success=False,
                message=f"获取表格信息失败: {str(e)}",
                error=str(e)
            )
    
    async def _search_spreadsheet(self, params: Dict[str, Any], is_excel: bool) -> SkillResult:
        """搜索表格数据"""
        file_path = params.get("file_path", "")
        search_value = params.get("search_value", "")
        sheet_name = params.get("sheet_name", "")
        
        if not search_value:
            return SkillResult(
                success=False,
                message="需要搜索值",
                error="Search value required"
            )
        
        if not os.path.exists(file_path):
            return SkillResult(
                success=False,
                message=f"文件不存在: {file_path}",
                error="File not found"
            )
        
        try:
            if is_excel:
                results = self._search_excel(file_path, search_value, sheet_name)
            else:
                results = self._search_csv(file_path, search_value)
            
            return SkillResult(
                success=True,
                message=f"搜索到 {len(results)} 个匹配项",
                data={
                    "file_path": file_path,
                    "search_value": search_value,
                    "results": results,
                    "total_matches": len(results)
                }
            )
        except Exception as e:
            return SkillResult(
                success=False,
                message=f"搜索表格失败: {str(e)}",
                error=str(e)
            )
    
    async def _filter_spreadsheet(self, params: Dict[str, Any], is_excel: bool) -> SkillResult:
        """过滤表格数据"""
        file_path = params.get("file_path", "")
        filter_column = params.get("filter_column", 0)
        filter_value = params.get("filter_value", "")
        sheet_name = params.get("sheet_name", "")
        
        if not filter_value:
            return SkillResult(
                success=False,
                message="需要过滤值",
                error="Filter value required"
            )
        
        if not os.path.exists(file_path):
            return SkillResult(
                success=False,
                message=f"文件不存在: {file_path}",
                error="File not found"
            )
        
        try:
            if is_excel:
                filtered_data = self._filter_excel(file_path, filter_column, filter_value, sheet_name)
            else:
                filtered_data = self._filter_csv(file_path, filter_column, filter_value)
            
            return SkillResult(
                success=True,
                message=f"过滤到 {len(filtered_data)} 行数据",
                data={
                    "file_path": file_path,
                    "filter_column": filter_column,
                    "filter_value": filter_value,
                    "filtered_data": filtered_data,
                    "rows_filtered": len(filtered_data)
                }
            )
        except Exception as e:
            return SkillResult(
                success=False,
                message=f"过滤表格失败: {str(e)}",
                error=str(e)
            )
    
    async def _sort_spreadsheet(self, params: Dict[str, Any], is_excel: bool) -> SkillResult:
        """排序表格数据"""
        file_path = params.get("file_path", "")
        sort_column = params.get("sort_column", 0)
        sort_ascending = params.get("sort_ascending", True)
        sheet_name = params.get("sheet_name", "")
        
        if not os.path.exists(file_path):
            return SkillResult(
                success=False,
                message=f"文件不存在: {file_path}",
                error="File not found"
            )
        
        try:
            if is_excel:
                sorted_data = self._sort_excel(file_path, sort_column, sort_ascending, sheet_name)
            else:
                sorted_data = self._sort_csv(file_path, sort_column, sort_ascending)
            
            return SkillResult(
                success=True,
                message=f"排序 {len(sorted_data)} 行数据",
                data={
                    "file_path": file_path,
                    "sort_column": sort_column,
                    "sort_ascending": sort_ascending,
                    "sorted_data": sorted_data,
                    "rows_sorted": len(sorted_data)
                }
            )
        except Exception as e:
            return SkillResult(
                success=False,
                message=f"排序表格失败: {str(e)}",
                error=str(e)
            )
    
    async def _convert_spreadsheet(self, params: Dict[str, Any], is_excel: bool) -> SkillResult:
        """转换表格格式"""
        file_path = params.get("file_path", "")
        output_path = params.get("output_path", "")
        target_format = params.get("format", "csv" if is_excel else "excel")
        sheet_name = params.get("sheet_name", "")
        
        if not os.path.exists(file_path):
            return SkillResult(
                success=False,
                message=f"文件不存在: {file_path}",
                error="File not found"
            )
        
        if not output_path:
            # 生成默认输出路径
            base_name = os.path.splitext(file_path)[0]
            if target_format == "csv":
                output_path = f"{base_name}.csv"
            elif target_format == "excel":
                output_path = f"{base_name}.xlsx"
            elif target_format == "json":
                output_path = f"{base_name}.json"
            else:
                return SkillResult(
                    success=False,
                    message=f"不支持的输出格式: {target_format}",
                    error=f"Unsupported output format: {target_format}"
                )
        
        try:
            # 读取数据
            if is_excel:
                data, _ = self._read_excel(file_path, sheet_name)
            else:
                data = self._read_csv(file_path)
            
            # 写入目标格式
            if target_format == "csv":
                self._write_csv(output_path, data)
            elif target_format == "excel":
                self._write_excel(output_path, data, sheet_name=sheet_name)
            elif target_format == "json":
                with open(output_path, 'w', encoding='utf-8') as f:
                    json.dump(data, f, ensure_ascii=False, indent=2)
            
            output_size = os.path.getsize(output_path)
            
            return SkillResult(
                success=True,
                message=f"转换成功: {file_path} -> {output_path}",
                data={
                    "source_path": file_path,
                    "output_path": output_path,
                    "source_format": "excel" if is_excel else "csv",
                    "target_format": target_format,
                    "rows_converted": len(data),
                    "output_size": output_size
                }
            )
        except Exception as e:
            return SkillResult(
                success=False,
                message=f"转换表格失败: {str(e)}",
                error=str(e)
            )
    
    async def _spreadsheet_stats(self, params: Dict[str, Any], is_excel: bool) -> SkillResult:
        """统计表格数据"""
        file_path = params.get("file_path", "")
        sheet_name = params.get("sheet_name", "")
        
        if not os.path.exists(file_path):
            return SkillResult(
                success=False,
                message=f"文件不存在: {file_path}",
                error="File not found"
            )
        
        try:
            if is_excel:
                data, _ = self._read_excel(file_path, sheet_name)
            else:
                data = self._read_csv(file_path)
            
            if not data:
                return SkillResult(
                    success=True,
                    message="表格为空",
                    data={"empty": True}
                )
            
            # 基本统计
            total_rows = len(data)
            total_columns = len(data[0]) if data else 0
            total_cells = total_rows * total_columns
            
            # 数据类型统计
            type_stats = {"text": 0, "number": 0, "date": 0, "empty": 0, "other": 0}
            
            for row in data:
                for cell in row:
                    if cell is None or cell == "":
                        type_stats["empty"] += 1
                    elif isinstance(cell, (int, float)):
                        type_stats["number"] += 1
                    elif isinstance(cell, str):
                        # 简单判断是否为日期
                        try:
                            datetime.fromisoformat(cell.replace('Z', '+00:00'))
                            type_stats["date"] += 1
                        except:
                            type_stats["text"] += 1
                    else:
                        type_stats["other"] += 1
            
            # 数值列统计
            numeric_stats = {}
            if total_columns > 0:
                for col_idx in range(total_columns):
                    col_values = []
                    for row in data:
                        if col_idx < len(row):
                            cell = row[col_idx]
                            if isinstance(cell, (int, float)):
                                col_values.append(cell)
                    
                    if col_values:
                        numeric_stats[col_idx] = {
                            "count": len(col_values),
                            "sum": sum(col_values),
                            "avg": sum(col_values) / len(col_values),
                            "min": min(col_values),
                            "max": max(col_values)
                        }
            
            return SkillResult(
                success=True,
                message=f"表格统计信息",
                data={
                    "file_path": file_path,
                    "total_rows": total_rows,
                    "total_columns": total_columns,
                    "total_cells": total_cells,
                    "type_stats": type_stats,
                    "numeric_stats": numeric_stats,
                    "has_data": total_cells > 0
                }
            )
        except Exception as e:
            return SkillResult(
                success=False,
                message=f"统计表格失败: {str(e)}",
                error=str(e)
            )
    
    # ===== Excel操作 =====
    
    def _read_excel(self, file_path: str, sheet_name: str = "", 
                   start_row: int = 0, end_row: int = -1,
                   start_col: int = 0, end_col: int = -1) -> Tuple[List[List[Any]], List[str]]:
        """读取Excel文件"""
        try:
            import openpyxl
            
            wb = openpyxl.load_workbook(file_path, data_only=True)
            sheet_names = wb.sheetnames
            
            if sheet_name and sheet_name in sheet_names:
                ws = wb[sheet_name]
            else:
                ws = wb.active
            
            data = []
            max_row = ws.max_row if end_row == -1 else min(end_row, ws.max_row)
            max_col = ws.max_column if end_col == -1 else min(end_col, ws.max_column)
            
            for row in ws.iter_rows(min_row=start_row + 1, max_row=max_row + 1,
                                   min_col=start_col + 1, max_col=max_col + 1):
                row_data = [cell.value for cell in row]
                data.append(row_data)
            
            return data, sheet_names
            
        except ImportError:
            logger.error("openpyxl未安装，无法处理Excel文件")
            raise ImportError("请安装openpyxl: pip install openpyxl")
    
    def _write_excel(self, file_path: str, data: List[List[Any]], 
                    headers: List[str] = None, sheet_name: str = "Sheet1"):
        """写入Excel文件"""
        try:
            import openpyxl
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name
            
            # 写入表头
            if headers:
                for col_idx, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col_idx, value=header)
                
                start_row = 2
            else:
                start_row = 1
            
            # 写入数据
            for row_idx, row_data in enumerate(data, start_row):
                for col_idx, cell_value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=cell_value)
            
            wb.save(file_path)
            
        except ImportError:
            logger.error("openpyxl未安装，无法处理Excel文件")
            raise ImportError("请安装openpyxl: pip install openpyxl")
    
    def _get_excel_info(self, file_path: str) -> Tuple[List[str], List[int], List[int]]:
        """获取Excel文件信息"""
        try:
            import openpyxl
            
            wb = openpyxl.load_workbook(file_path, read_only=True)
            sheet_names = wb.sheetnames
            row_counts = []
            column_counts = []
            
            for sheet_name in sheet_names:
                ws = wb[sheet_name]
                row_counts.append(ws.max_row)
                column_counts.append(ws.max_column)
            
            return sheet_names, row_counts, column_counts
            
        except ImportError:
            logger.error("openpyxl未安装，无法处理Excel文件")
            raise ImportError("请安装openpyxl: pip install openpyxl")
    
    def _search_excel(self, file_path: str, search_value: str, sheet_name: str = "") -> List[Dict[str, Any]]:
        """搜索Excel文件"""
        data, sheet_names = self._read_excel(file_path, sheet_name)
        results = []
        
        for row_idx, row in enumerate(data):
            for col_idx, cell in enumerate(row):
                if cell and search_value.lower() in str(cell).lower():
                    results.append({
                        "sheet": sheet_name or sheet_names[0],
                        "row": row_idx + 1,
                        "column": col_idx + 1,
                        "value": str(cell)[:100]
                    })
        
        return results
    
    def _filter_excel(self, file_path: str, filter_column: int, filter_value: str, 
                     sheet_name: str = "") -> List[List[Any]]:
        """过滤Excel数据"""
        data, _ = self._read_excel(file_path, sheet_name)
        
        if not data:
            return []
        
        filtered_data = []
        for row in data:
            if filter_column < len(row):
                cell_value = row[filter_column]
                if cell_value and filter_value.lower() in str(cell_value).lower():
                    filtered_data.append(row)
        
        return filtered_data
    
    def _sort_excel(self, file_path: str, sort_column: int, ascending: bool = True,
                   sheet_name: str = "") -> List[List[Any]]:
        """排序Excel数据"""
        data, _ = self._read_excel(file_path, sheet_name)
        
        if not data:
            return []
        
        # 检查是否所有行的指定列都有值
        sortable_data = []
        for row in data:
            if sort_column < len(row) and row[sort_column] is not None:
                sortable_data.append(row)
        
        # 排序
        try:
            sorted_data = sorted(sortable_data, 
                               key=lambda x: x[sort_column], 
                               reverse=not ascending)
        except:
            # 如果排序失败，按字符串排序
            sorted_data = sorted(sortable_data,
                               key=lambda x: str(x[sort_column]) if x[sort_column] is not None else "",
                               reverse=not ascending)
        
        return sorted_data
    
    # ===== CSV操作 =====
    
    def _read_csv(self, file_path: str, start_row: int = 0, end_row: int = -1,
                 start_col: int = 0, end_col: int = -1) -> List[List[Any]]:
        """读取CSV文件"""
        delimiter = ','
        quotechar = '"'
        
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                reader = csv.reader(f, delimiter=delimiter, quotechar=quotechar)
                data = list(reader)
            
            # 应用行列限制
            if end_row == -1:
                end_row = len(data)
            
            filtered_data = []
            for row_idx in range(start_row, min(end_row, len(data))):
                row = data[row_idx]
                if end_col == -1:
                    end_col = len(row)
                filtered_row = row[start_col:end_col]
                filtered_data.append(filtered_row)
            
            return filtered_data
        except Exception as e:
            logger.error(f"读取CSV失败: {e}")
            return []
    
    def _write_csv(self, file_path: str, data: List[List[Any]], headers: List[str] = None):
        """写入CSV文件"""
        with open(file_path, 'w', encoding='utf-8', newline='') as f:
            writer = csv.writer(f)
            
            if headers:
                writer.writerow(headers)
            
            writer.writerows(data)
    
    def _get_csv_info(self, file_path: str) -> Tuple[List[str], List[int], List[int]]:
        """获取CSV文件信息"""
        data = self._read_csv(file_path)
        
        if not data:
            return ["CSV Data"], [0], [0]
        
        return ["CSV Data"], [len(data)], [len(data[0]) if data else 0]
    
    def _search_csv(self, file_path: str, search_value: str) -> List[Dict[str, Any]]:
        """搜索CSV文件"""
        data = self._read_csv(file_path)
        results = []
        
        for row_idx, row in enumerate(data):
            for col_idx, cell in enumerate(row):
                if cell and search_value.lower() in str(cell).lower():
                    results.append({
                        "sheet": "CSV Data",
                        "row": row_idx + 1,
                        "column": col_idx + 1,
                        "value": str(cell)[:100]
                    })
        
        return results
    
    def _filter_csv(self, file_path: str, filter_column: int, filter_value: str) -> List[List[Any]]:
        """过滤CSV数据"""
        data = self._read_csv(file_path)
        
        if not data:
            return []
        
        filtered_data = []
        for row in data:
            if filter_column < len(row):
                cell_value = row[filter_column]
                if cell_value and filter_value.lower() in str(cell_value).lower():
                    filtered_data.append(row)
        
        return filtered_data
    
    def _sort_csv(self, file_path: str, sort_column: int, ascending: bool = True) -> List[List[Any]]:
        """排序CSV数据"""
        data = self._read_csv(file_path)
        
        if not data:
            return []
        
        # 检查是否所有行的指定列都有值
        sortable_data = []
        for row in data:
            if sort_column < len(row) and row[sort_column]:
                sortable_data.append(row)
        
        # 排序
        try:
            sorted_data = sorted(sortable_data, 
                               key=lambda x: x[sort_column], 
                               reverse=not ascending)
        except:
            # 如果排序失败，按字符串排序
            sorted_data = sorted(sortable_data,
                               key=lambda x: str(x[sort_column]) if x[sort_column] else "",
                               reverse=not ascending)
        
        return sorted_data
    
    def _format_size(self, size_bytes: int) -> str:
        """格式化文件大小"""
        if size_bytes == 0:
            return "0 B"
        
        units = ["B", "KB", "MB", "GB", "TB"]
        unit_index = 0
        
        size = float(size_bytes)
        while size >= 1024 and unit_index < len(units) - 1:
            size /= 1024
            unit_index += 1
        
        return f"{size:.2f} {units[unit_index]}"